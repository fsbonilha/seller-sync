from datetime import datetime
import logging
import pandas as pd
import openpyxl as xl
import os
import re
from tqdm import tqdm
import win32com.client as win32

logging.basicConfig(filename='app.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


class SpreadsheetSplitter:
    CONTACTS_SHEET = "CONTATOS"

    def __init__(self, template_file: str, input_data_file: str,
                 input_sheet_names: list, output_folder: str, id_column: str,
                 filename_location: dict) -> None:
        self.template_file = template_file
        self.input_data_file = input_data_file
        self.input_sheet_names = input_sheet_names
        self.output_folder = output_folder
        self.template_file = self.load_spreadsheet()
        self.dfs = self.import_data()
        self.files = []
        self.id_column = id_column
        self.filename_location = filename_location
        self.contacts = self.get_contacts()

    def import_data(self) -> dict:
        """
        Returns a dict with sheet names as keys
        """
        logging.info("Importing data")

        dfs = pd.read_excel(self.input_data_file, sheet_name=None, header=0)

        logging.info(f"Sheets imported: {dfs.keys()}")
        return dfs

    def split_sellers(self) -> None:
        """
        Create a spreadsheet based on `template_file` for each seller
        """
        logging.info("Started splitting merchants")
        merchant_list = list(map(int, self.contacts[self.id_column]))

        for merchant_id in tqdm(merchant_list, desc="Splitting Files"):
            logging.info(f"Splitting {merchant_id}...")
            merchant_dfs = dict()
            for sheet_name in self.input_sheet_names:
                df = self.dfs[sheet_name]
                df = df[df[self.id_column] == merchant_id]
                merchant_dfs[sheet_name] = df
            self.clean_spreadsheet()
            self.fill_cols(merchant_dfs)
            self.save_spreadsheet(merchant_dfs)

    def fill_cols(self, merchant_dfs: dict) -> None:
        logging.info("Filling spreadsheet")
        for ws in self.input_sheet_names:
            worksheet = self.template_file[ws]
            current_df = merchant_dfs[ws]
            cols = current_df.columns

            for ncol, col in enumerate(cols):
                worksheet.cell(row=1, column=ncol + 1, value=col)
            for index, row in current_df.reset_index().iterrows():
                row_index = index + 2  # considering excel 1 based indexes
                for ncol, col in enumerate(current_df.columns):
                    worksheet.cell(row=row_index, column=ncol + 1,
                                   value=row[col])
        logging.info("Spreadsheet filled")

    def get_save_path(self, merchant_dfs) -> str:
        filename_sheet = merchant_dfs[self.filename_location["sheet"]]
        seller_name = (
            filename_sheet[self.filename_location["col"]]
            .iloc[0]
            .replace(' ', '_')
        )

        clean_seller_name = re.sub('[^A-Za-z0-9]+', '', seller_name)
        path = f"{self.output_folder}/relatorio-{clean_seller_name}.xlsx"

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
        return path

    def load_spreadsheet(self) -> xl.Workbook:
        logging.info("Loading Spreadsheet")
        return xl.load_workbook(self.template_file)

    def clean_spreadsheet(self) -> None:  # Preparing for next fill_cols
        logging.info("Cleaning spreadsheet")

        for ws in self.input_sheet_names:
            for row in self.template_file[ws].iter_rows(min_row=1, min_col=1, max_row=50, max_col=50):
                for cell in row:
                    cell.value = None

    def save_spreadsheet(self, merchant_dfs: dict) -> None:
        logging.info("Saving Spreadsheet")

        path = self.get_save_path(merchant_dfs)

        self.template_file.save(path)
        self.template_file = xl.load_workbook(path)   # Open saved file
        self.files.append(path)

        logging.info(f"File saved to {path}")

    def refresh_pivot(self) -> None:
        logging.info("Refreshing pivot")
        worksheet = self.template_file["Pivot"]
        pivot = worksheet._pivots[0]
        pivot.cache.refreshOnLoad = True

    def get_contacts(self) -> pd.DataFrame:
        """
        Returns a Dataframe with columns: email, merchant_customer_id, subject, body
        """
        contacts_df = self.dfs[self.CONTACTS_SHEET]
        return pd.DataFrame(
            contacts_df.values[1:],
            columns=contacts_df.iloc[0]
        )


class Mailman:
    def __init__(self, recipient, subject, body, attachment):
        self.email_list = recipient
        self.subject = subject
        self.body = body
        self.attachment = attachment

    def send_email(self):
        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = self.email_list
            mail.Subject = self.subject
            mail.Body = self.body
            # mail.HTMLBody = '<h2>HTML here</h2>'
            attachment = self.attachment
            mail.Attachments.Add(attachment)

            mail.Send()
        except Exception as e:
            print(f"Error ocurred while sending email: {str(e)}")


def send_multiple_emails(files, emails, subject, body):
    assert len(files) == len(emails), \
        "Len of File List is not equal Len of Email List"

    for idx, file in tqdm(enumerate(files), desc="Sending Emails"):
        mail_man = Mailman(
            recipient=emails[idx],
            subject=subject,
            body=body,
            attachment=os.path.abspath(file)
        )
        mail_man.send_email()


def confirm_send_email():
    return (input(f"Please check files and type '{CONFIRMATION_KEYWORD}' to send emails: ") == CONFIRMATION_KEYWORD)


CONFIRMATION_KEYWORD = "send"
COLUMN_EMAIL = "email"
COLUMN_SUBJECT = "email_subject"
COLUMN_BODY = "email_body"
COLUMN_ID = "merchant_customer_id"
FILE_TEMPLATE = "template.xlsx"
FILE_DATA = "SellerSync_Data.xlsx"
SHEETS_INPUT = ["GMS_AGG", "GMS_SKU"]
SHEETS_FILENAME = {"col": "seller_name", "sheet": "GMS_AGG"}
OUTPUT_FOLDER = "output"


def main():
    deal_splitter = SpreadsheetSplitter(
        template_file=FILE_TEMPLATE,
        input_data_file=FILE_DATA,
        input_sheet_names=SHEETS_INPUT,
        output_folder=OUTPUT_FOLDER,
        id_column=COLUMN_ID,
        filename_location=SHEETS_FILENAME
    )
    deal_splitter.split_sellers()

    print(deal_splitter.contacts)

    if confirm_send_email():
        send_multiple_emails(
            files=deal_splitter.files,
            emails=deal_splitter.contacts[COLUMN_EMAIL],
            subject=deal_splitter.contacts[COLUMN_SUBJECT][0],
            body=deal_splitter.contacts[COLUMN_BODY][0]
        )


if __name__ == "__main__":
    main()
